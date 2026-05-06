# -*- coding: utf-8 -*-
"""
本地 xlsx 讀寫客戶端（模擬 sheets_client.py 的 API）。

職責：
1. 從本地 xlsx 讀取待處理的新聞列（A~E 欄）
2. 將爬取結果寫回本地 xlsx 的 F~L 欄位

所有 API 與 sheets_client.py 保持一致，方便切換。
"""
from __future__ import annotations

import os
import openpyxl
from datetime import datetime, timezone, timedelta
from typing import Optional

import config

_TZ_TAIPEI = timezone(timedelta(hours=8))

# xlsx 欄位對應（1-based index），與 config.COL 一致
_COL = {
    "日期":       1,   # A
    "標題":       2,   # B
    "短網址":     3,   # C
    "新聞網址":   4,   # D
    "真實網址":   5,   # E
    "狀態":       6,   # F
    "內文":       7,   # G
    "擷取方法":   8,   # H
    "診斷資訊":   9,   # I
    "最後嘗試":  10,   # J
    "字數":      11,   # K
    "網域":      12,   # L
}

# 狀態值
STATUS_PENDING = "PENDING"
STATUS_DONE = "DONE"
STATUS_FAILED = "FAILED"

ACTIONABLE_STATUSES = {STATUS_PENDING, "RETRY_1", "RETRY_2", "RETRY_3"}


class LocalWorkbook:
    """包裝 openpyxl，模擬 gspread Worksheet 行為。"""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath)
        self.ws = self.wb.active
        self._all_values: list[list] | None = None

    def get_all_values(self) -> list[list]:
        """快取讀取，避免重複磁碟 IO。"""
        if self._all_values is None:
            self._all_values = list(self.ws.values)
        return self._all_values

    def reload(self):
        """重新從磁碟讀取。"""
        self.wb = openpyxl.load_workbook(self.filepath)
        self.ws = self.wb.active
        self._all_values = None

    def save(self):
        """寫回磁碟。"""
        self._all_values = None  # 清除快取
        self.wb.save(self.filepath)


# 全域 workbook 實例（lazy init）
_workbook: LocalWorkbook | None = None
_xlsx_path: str | None = None


def init(xlsx_path: str):
    """初始化本地 xlsx 檔案路徑。"""
    global _workbook, _xlsx_path
    _xlsx_path = xlsx_path
    _workbook = LocalWorkbook(xlsx_path)
    _ensure_headers(_workbook)


def _ensure_headers(wb: LocalWorkbook) -> None:
    """確保 F~L 標題存在，否則寫入。"""
    ws = wb.ws
    headers = {6: "狀態", 7: "內文", 8: "擷取方法", 9: "診斷資訊", 10: "最後嘗試", 11: "字數", 12: "網域"}
    needs_save = False
    for col, label in headers.items():
        existing = ws.cell(row=1, column=col).value
        if not existing:
            ws.cell(row=1, column=col, value=label)
            needs_save = True
    if needs_save:
        wb.save()


def _get_ws() -> LocalWorkbook:
    if _workbook is None:
        raise RuntimeError("請先呼叫 init('/path/to/news_trimmed.xlsx')")
    return _workbook


# =============================================================================
# 讀取
# =============================================================================

def get_pending_rows(batch_size: int = config.BATCH_SIZE,
                     status_col: int = _COL["狀態"],
                     url_col: int = _COL["真實網址"],
                     title_col: int = _COL["標題"],
                     date_col: int = _COL["日期"]) -> list[dict]:
    """
    讀取需要處理的列，優先順序：PENDING > RETRY_1 > RETRY_2 > RETRY_3。

    回傳 list of dict，每個 dict 包含:
      - row_index: xlsx 中的列號 (1-based)
      - url: 真實網址
      - title: 新聞標題
      - status: 目前狀態
      - date: 新聞日期
    """
    wb = _get_ws()
    all_values = wb.get_all_values()

    if len(all_values) <= 1:
        return []

    rows = []
    for i, row in enumerate(all_values[1:], start=2):  # 跳過 header，列號從 2 開始
        status = row[status_col - 1] if len(row) >= status_col else ""

        # 如果狀態欄為空或不在 ACTIONABLE_STATUSES 中，跳過（視為已處理）
        if status and status not in ACTIONABLE_STATUSES:
            continue

        url = row[url_col - 1] if len(row) >= url_col else ""
        title = row[title_col - 1] if len(row) >= title_col else ""
        date = row[date_col - 1] if len(row) >= date_col else ""

        if not url:
            continue

        rows.append({
            "row_index": i,
            "url": url,
            "title": title,
            "status": status or STATUS_PENDING,  # 空值預設 PENDING
            "date": date,
        })

    # 排序：PENDING 優先，然後 RETRY_1 > RETRY_2 > RETRY_3
    priority = {STATUS_PENDING: 0, "RETRY_1": 1, "RETRY_2": 2, "RETRY_3": 3}
    rows.sort(key=lambda r: priority.get(r["status"], 99))

    return rows[:batch_size]


def get_sheet_stats(status_col: int = _COL["狀態"]) -> dict:
    """取得 Sheet 的狀態統計，用於報告。"""
    wb = _get_ws()
    all_values = wb.get_all_values()

    stats = {"total": len(all_values) - 1}
    for row in all_values[1:]:
        status = row[status_col - 1] if len(row) >= status_col else "UNKNOWN"
        stats[status] = stats.get(status, 0) + 1

    return stats


# =============================================================================
# 寫入
# =============================================================================

def _truncate_if_needed(field_name: str, value: str) -> str:
    """對內文欄位截斷，避免超出 Google Sheets 上限。"""
    if field_name == "內文" and len(value) > config.MAX_CONTENT_LENGTH:
        return value[:config.MAX_CONTENT_LENGTH]
    return value


def update_row(row_index: int, fields: dict) -> None:
    """
    更新 xlsx 中指定列的多個欄位（F~L）。

    fields 的 key 對應 _COL 的欄位名稱。
    """
    wb = _get_ws()
    ws = wb.ws

    for field_name, value in fields.items():
        col_index = _COL.get(field_name)
        if col_index is None:
            continue
        str_value = str(value) if value else ""
        str_value = _truncate_if_needed(field_name, str_value)
        ws.cell(row=row_index, column=col_index, value=str_value)


def batch_update_rows(updates: list[tuple[int, dict]]) -> None:
    """
    批次更新多列。updates = [(row_index, fields_dict), ...]

    為減少磁碟 IO，一次寫入所有 cell 後再儲存。
    """
    wb = _get_ws()
    ws = wb.ws

    for row_index, fields in updates:
        for field_name, value in fields.items():
            col_index = _COL.get(field_name)
            if col_index is None:
                continue
            str_value = str(value) if value else ""
            str_value = _truncate_if_needed(field_name, str_value)
            ws.cell(row=row_index, column=col_index, value=str_value)

    wb.save()


def flush() -> None:
    """將目前所有改動寫回磁碟。"""
    wb = _get_ws()
    wb.save()
